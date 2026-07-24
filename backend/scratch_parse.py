import openpyxl

def run():
    file_path = '테트리스U7_240808_decrypted.xlsm'
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print("Sheets:", wb.sheetnames)
    except Exception as e:
        print("Error loading workbook:", e)
        return

    def print_sheet(sheet_name, max_row=40, max_col=15):
        if sheet_name in wb.sheetnames:
            print(f"--- Sheet: {sheet_name} (Formulas) ---")
            ws = wb[sheet_name]
            for r in range(1, max_row):
                row_data = []
                for c in range(1, max_col):
                    val = ws.cell(row=r, column=c).value
                    row_data.append(str(val) if val is not None else "")
                if any(row_data):
                    print(f"Row {r}: {row_data}")
        else:
            print(f"Sheet {sheet_name} not found.")

    print_sheet('DDM', max_row=100, max_col=20)
    print_sheet('테트리스', max_row=50, max_col=20)

if __name__ == '__main__':
    run()
